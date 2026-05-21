#!/usr/bin/env python3
"""
Follow-up Sender — L&D Designs
Finds leads that need a follow-up (3 days with no reply) and sends them.

Usage:
    python send_followups.py --dry-run
    python send_followups.py --limit 10
"""

import subprocess, sys, argparse, json
from pathlib import Path
from datetime import datetime, timedelta

def _ensure(pkg, import_as=None):
    try:
        __import__(import_as or pkg)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", pkg])

_ensure("openpyxl")

import openpyxl
from openpyxl.styles import PatternFill

sys.path.insert(0, str(Path(__file__).parent.parent / "outreach"))
from templates import get_followup_email, get_followup_whatsapp

CRM_FILE = Path(__file__).parent.parent / "leads" / "master_crm.xlsx"
LOG_FILE = Path(__file__).parent / "followup_log.json"
FOLLOWUP_DAYS = 3
MAX_FOLLOWUPS = 2

COLS = {
    "name":         1,
    "industry":     2,
    "city":         3,
    "email":        5,
    "phone":        6,
    "first_contact":11,
    "followup_sent":12,
    "replied":      13,
    "last_contact": 18,
    "next_followup":19,
    "status":       20,
}


def load_log():
    if LOG_FILE.exists():
        try:
            return json.loads(LOG_FILE.read_text())
        except Exception:
            pass
    return {"followups": [], "total_sent": 0}

def save_log(log):
    LOG_FILE.write_text(json.dumps(log, indent=2))

def needs_followup(row):
    status = row[COLS["status"] - 1].value
    replied = row[COLS["replied"] - 1].value
    next_date_cell = row[COLS["next_followup"] - 1].value

    if status not in ("Contacted", "Follow-up 1"):
        return False
    if replied:
        return False
    if not next_date_cell:
        return False

    try:
        if isinstance(next_date_cell, str):
            next_date = datetime.strptime(next_date_cell, "%d/%m/%Y")
        else:
            next_date = next_date_cell
        return datetime.now() >= next_date
    except Exception:
        return False

def get_followup_step(status):
    if status == "Contacted":
        return 1
    elif status == "Follow-up 1":
        return 2
    return None

def update_crm_after_followup(ws, row_num, step):
    status_map = {1: "Follow-up 1", 2: "Follow-up 2"}
    colour_map = {1: "FFF2CC", 2: "FFE0B2"}
    new_status = status_map.get(step, "Follow-up 2")
    ws.cell(row_num, COLS["status"]).value = new_status
    ws.cell(row_num, COLS["status"]).fill  = PatternFill("solid", fgColor=colour_map.get(step, "FFE0B2"))
    ws.cell(row_num, COLS["followup_sent"]).value = datetime.now().strftime("%d/%m/%Y")
    ws.cell(row_num, COLS["last_contact"]).value  = datetime.now().strftime("%d/%m/%Y")
    if step < MAX_FOLLOWUPS:
        next_date = datetime.now() + timedelta(days=FOLLOWUP_DAYS)
        ws.cell(row_num, COLS["next_followup"]).value = next_date.strftime("%d/%m/%Y")
    else:
        ws.cell(row_num, COLS["next_followup"]).value = ""

def send_followup_email(to_email, subject, body):
    import smtplib, os
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    gmail_addr = os.getenv("GMAIL_ADDRESS")
    gmail_pass = os.getenv("GMAIL_APP_PASSWORD")

    if not gmail_addr or not gmail_pass:
        raise EnvironmentError("Set GMAIL_ADDRESS and GMAIL_APP_PASSWORD")

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = gmail_addr
    msg["To"]      = to_email
    msg.attach(MIMEText(body, "plain"))

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(gmail_addr, gmail_pass)
        server.sendmail(gmail_addr, to_email, msg.as_string())

def run(dry_run, limit):
    if not CRM_FILE.exists():
        print("  CRM not found. Run: python BusinessOS/leads/create_crm.py")
        return

    wb = openpyxl.load_workbook(CRM_FILE)
    ws = wb["CRM"]
    log = load_log()

    due = []
    for row in ws.iter_rows(min_row=2):
        name = row[COLS["name"] - 1].value
        if not name:
            continue
        if needs_followup(row):
            step = get_followup_step(str(row[COLS["status"] - 1].value or ""))
            if step and step <= MAX_FOLLOWUPS:
                due.append({
                    "row":   row[0].row,
                    "name":  str(name),
                    "email": str(row[COLS["email"] - 1].value or ""),
                    "phone": str(row[COLS["phone"] - 1].value or ""),
                    "step":  step,
                })
        if len(due) >= limit:
            break

    print(f"\n  {len(due)} leads due for follow-up\n")

    if dry_run:
        for lead in due:
            f = get_followup_email(lead["step"], lead["name"])
            print(f"  [{lead['name']}] Step {lead['step']} → {lead['email'] or 'no email'}")
            print(f"  Subject: {f['subject']}\n")
        print("  Dry run — no messages sent.\n")
        return

    sent = 0
    for lead in due:
        if sent >= limit:
            break
        try:
            if lead["email"]:
                f = get_followup_email(lead["step"], lead["name"])
                send_followup_email(lead["email"], f["subject"], f["body"])
                update_crm_after_followup(ws, lead["row"], lead["step"])
                log["followups"].append({
                    "business": lead["name"],
                    "step":     lead["step"],
                    "to":       lead["email"],
                    "date":     datetime.now().strftime("%d/%m/%Y %H:%M"),
                })
                sent += 1
                print(f"  Follow-up {lead['step']} sent → {lead['name']} ({lead['email']})")
            else:
                wa = get_followup_whatsapp(lead["step"], lead["name"])
                update_crm_after_followup(ws, lead["row"], lead["step"])
                log["followups"].append({
                    "business": lead["name"],
                    "step":     lead["step"],
                    "channel":  "whatsapp",
                    "message":  wa,
                    "date":     datetime.now().strftime("%d/%m/%Y %H:%M"),
                })
                sent += 1
                print(f"  Follow-up {lead['step']} logged → {lead['name']}")
        except Exception as e:
            print(f"  Failed → {lead['name']}: {e}")
        import time
        time.sleep(3)

    log["total_sent"] = log.get("total_sent", 0) + sent
    log["last_run"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    save_log(log)
    wb.save(CRM_FILE)
    print(f"\n  Done. {sent} follow-ups sent.\n")

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--limit",   type=int, default=20)
    args = parser.parse_args()

    print("\n  L&D Designs — Follow-up Sender")
    run(args.dry_run, args.limit)

if __name__ == "__main__":
    main()
