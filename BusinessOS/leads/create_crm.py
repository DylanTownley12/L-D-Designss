#!/usr/bin/env python3
"""
Master CRM Generator — L&D Designs
Creates or updates the master CRM spreadsheet.

Usage:
    python create_crm.py                    # Create fresh CRM
    python create_crm.py --import leads.xlsx  # Import leads from a leads file
"""

import subprocess, sys, argparse
from pathlib import Path

def _ensure(pkg, import_as=None):
    try:
        __import__(import_as or pkg)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", pkg])

_ensure("openpyxl")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

CRM_FILENAME = Path(__file__).parent / "master_crm.xlsx"

CRM_COLUMNS = [
    ("Business Name",     28),
    ("Industry",          22),
    ("City",              14),
    ("Contact Name",      20),
    ("Email",             34),
    ("Phone",             18),
    ("Website Status",    18),
    ("Facebook Link",     34),
    ("Instagram Link",    34),
    ("Date Added",        14),
    ("First Contact Sent",16),
    ("Follow-up Sent",    16),
    ("Replied",            10),
    ("Interested",         12),
    ("Not Interested",     14),
    ("Booked Call",        12),
    ("Notes",              36),
    ("Last Contact Date",  16),
    ("Next Follow-up Date",18),
    ("Outreach Status",    18),
]

STATUS_OPTIONS = [
    "Not Contacted",
    "Contacted",
    "Follow-up 1",
    "Follow-up 2",
    "Replied",
    "Interested",
    "Closed",
    "No Response",
]

STATUS_COLOURS = {
    "Not Contacted": "E8E8E8",
    "Contacted":     "D6EAFF",
    "Follow-up 1":   "FFF2CC",
    "Follow-up 2":   "FFE0B2",
    "Replied":       "E1F5E1",
    "Interested":    "C8F7C5",
    "Closed":        "D4EDDA",
    "No Response":   "F8D7DA",
}

FILL_H = PatternFill("solid", fgColor="1A2035")
THIN   = Border(**{s: Side(style="thin", color="CCCCCC") for s in ("left", "right", "top", "bottom")})


def _status_fill(status):
    colour = STATUS_COLOURS.get(status, "FFFFFF")
    return PatternFill("solid", fgColor=colour)


def create_crm(leads=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CRM"

    for col, (hdr, w) in enumerate(CRM_COLUMNS, 1):
        c = ws.cell(1, col, hdr)
        c.fill = FILL_H
        c.font = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 32

    start_row = 2
    if leads:
        for ri, lead in enumerate(leads, start_row):
            status = "Not Contacted"
            fill = _status_fill(status)
            row_vals = [
                lead.get("Business Name") or lead.get("name", ""),
                lead.get("Industry") or lead.get("business_type", ""),
                lead.get("City", ""),
                lead.get("Contact Name", ""),
                lead.get("Email") or lead.get("email", ""),
                lead.get("Phone") or lead.get("phone", ""),
                lead.get("Website Status") or lead.get("website_status", ""),
                lead.get("Facebook Link", ""),
                lead.get("Instagram Link", ""),
                datetime.now().strftime("%d/%m/%Y"),
                "",  # First Contact Sent
                "",  # Follow-up Sent
                "",  # Replied
                "",  # Interested
                "",  # Not Interested
                "",  # Booked Call
                lead.get("Notes") or lead.get("status_notes", ""),
                "",  # Last Contact Date
                "",  # Next Follow-up Date
                status,
            ]
            for col, val in enumerate(row_vals, 1):
                c = ws.cell(ri, col, val)
                c.fill = fill
                c.alignment = Alignment(vertical="center")
                c.border = THIN
            ws.row_dimensions[ri].height = 18
        start_row = len(leads) + 2

    # Add 50 empty rows for manual entries
    for ri in range(start_row, start_row + 50):
        for col in range(1, len(CRM_COLUMNS) + 1):
            c = ws.cell(ri, col, "")
            c.fill = PatternFill("solid", fgColor="FAFAFA") if ri % 2 == 0 else PatternFill(fill_type=None)
            c.alignment = Alignment(vertical="center")
            c.border = THIN
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(CRM_COLUMNS))}1"

    # Status key sheet
    ws2 = wb.create_sheet("Status Guide")
    ws2.cell(1, 1, "Status").font = Font(bold=True, size=11)
    ws2.cell(1, 2, "Meaning").font = Font(bold=True, size=11)
    guide = [
        ("Not Contacted", "Lead found, haven't reached out yet"),
        ("Contacted",     "Sent first message (email/WhatsApp/phone)"),
        ("Follow-up 1",   "Sent first follow-up after no reply"),
        ("Follow-up 2",   "Sent second follow-up"),
        ("Replied",       "They replied — conversation started"),
        ("Interested",    "Actively interested in a website"),
        ("Closed",        "Booked and job done"),
        ("No Response",   "No reply after multiple attempts"),
    ]
    for r, (status, meaning) in enumerate(guide, 2):
        c1 = ws2.cell(r, 1, status)
        c1.fill = _status_fill(status)
        c1.font = Font(bold=True, size=10)
        c1.alignment = Alignment(vertical="center")
        ws2.cell(r, 2, meaning).alignment = Alignment(vertical="center")
        ws2.row_dimensions[r].height = 20
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 46

    wb.save(CRM_FILENAME)
    print(f"  CRM saved: {CRM_FILENAME}")
    return CRM_FILENAME


def import_leads_from_file(leads_file):
    try:
        wb = openpyxl.load_workbook(leads_file)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        leads = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                leads.append(dict(zip(headers, row)))
        print(f"  Loaded {len(leads)} leads from {leads_file}")
        return leads
    except Exception as e:
        print(f"  Could not load {leads_file}: {e}")
        return []


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--import", dest="import_file", help="Import leads from a file")
    args = parser.parse_args()

    leads = []
    if args.import_file:
        p = Path(args.import_file)
        if not p.exists():
            p = Path(__file__).parent / args.import_file
        if p.exists():
            leads = import_leads_from_file(p)
        else:
            print(f"  File not found: {args.import_file}")

    print(f"\n  Creating master CRM with {len(leads)} leads...")
    create_crm(leads)
    print(f"  Done!\n")


if __name__ == "__main__":
    main()
