#!/usr/bin/env python3
"""
UK Lead Finder — L&D Designs
Finds local UK businesses with no website or weak online presence.

Usage:
    python find_leads_uk.py
    python find_leads_uk.py --city "Manchester" --radius 20
    python find_leads_uk.py --city "Birmingham" --types "Barbers,Plumbers"
"""

import subprocess, sys, argparse
from pathlib import Path

def _ensure(pkg, import_as=None):
    try:
        __import__(import_as or pkg)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", pkg])

for _p in [("requests",), ("openpyxl",), ("beautifulsoup4", "bs4"), ("lxml",)]:
    _ensure(*_p)

import re, math, time, unicodedata, json
from datetime import datetime
from urllib.parse import urljoin, quote

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

# ── UK City coordinates ───────────────────────────────────────────────────────
UK_CITIES = {
    "wigan":          (53.5450, -2.6325),
    "manchester":     (53.4808, -2.2426),
    "liverpool":      (53.4084, -2.9916),
    "birmingham":     (52.4862, -1.8904),
    "leeds":          (53.8008, -1.5491),
    "sheffield":      (53.3811, -1.4701),
    "london":         (51.5074, -0.1278),
    "bristol":        (51.4545, -2.5879),
    "nottingham":     (52.9548, -1.1581),
    "leicester":      (52.6369, -1.1398),
    "coventry":       (52.4068, -1.5197),
    "bradford":       (53.7960, -1.7594),
    "hull":           (53.7457, -0.3367),
    "newcastle":      (54.9783, -1.6178),
    "sunderland":     (54.9069, -1.3838),
    "bolton":         (53.5780, -2.4282),
    "stockport":      (53.4083, -2.1494),
    "salford":        (53.4875, -2.2901),
    "oldham":         (53.5409, -2.1114),
    "rochdale":       (53.6136, -2.1610),
    "bury":           (53.5935, -2.2986),
    "warrington":     (53.3900, -2.5970),
    "st helens":      (53.4540, -2.7360),
    "blackpool":      (53.8175, -3.0357),
    "preston":        (53.7632, -2.7031),
    "blackburn":      (53.7482, -2.4851),
    "burnley":        (53.7890, -2.2408),
    "stoke":          (53.0027, -2.1794),
    "wolverhampton":  (52.5868, -2.1257),
    "derby":          (52.9225, -1.4746),
    "oxford":         (51.7520, -1.2577),
    "cambridge":      (52.2053, 0.1218),
    "portsmouth":     (50.8198, -1.0880),
    "southampton":    (50.9097, -1.4044),
    "reading":        (51.4543, -0.9781),
    "plymouth":       (50.3755, -4.1427),
    "exeter":         (50.7184, -3.5339),
    "cardiff":        (51.4816, -3.1791),
    "swansea":        (51.6214, -3.9436),
    "edinburgh":      (55.9533, -3.1883),
    "glasgow":        (55.8642, -4.2518),
    "belfast":        (54.5973, -5.9301),
}

# ── Business type presets ─────────────────────────────────────────────────────
PRESETS = {
    "Barbers & Hair Salons": {
        "yell": ["barbers", "hairdressers", "hair salon", "barbershop"],
        "osm":  [("shop", "hairdresser"), ("shop", "barber"), ("amenity", "hairdresser")],
    },
    "Plumbers": {
        "yell": ["plumbers", "plumbing services", "emergency plumber"],
        "osm":  [("craft", "plumber")],
    },
    "Electricians": {
        "yell": ["electricians", "electrical contractor", "emergency electrician"],
        "osm":  [("craft", "electrician")],
    },
    "Roofers": {
        "yell": ["roofers", "roofing contractors", "roof repairs"],
        "osm":  [("craft", "roofer")],
    },
    "Builders": {
        "yell": ["builders", "building contractors", "construction company"],
        "osm":  [("craft", "builder")],
    },
    "Mechanics & Garages": {
        "yell": ["car garage", "MOT centre", "car repair", "mechanic", "tyre fitting"],
        "osm":  [("shop", "car_repair"), ("shop", "tyres")],
    },
    "Painters & Decorators": {
        "yell": ["painters decorators", "painting decorating", "decorator"],
        "osm":  [("craft", "painter")],
    },
    "Gardeners & Landscapers": {
        "yell": ["landscaper", "gardener", "garden maintenance", "tree surgeon"],
        "osm":  [("craft", "gardener"), ("craft", "landscaper")],
    },
    "Restaurants & Takeaways": {
        "yell": ["restaurant", "takeaway", "cafe", "fish and chips", "pizza", "curry house"],
        "osm":  [("amenity", "restaurant"), ("amenity", "fast_food"), ("amenity", "cafe")],
    },
    "Beauty Salons": {
        "yell": ["beauty salon", "nail salon", "nail bar", "lashes", "aesthetics"],
        "osm":  [("shop", "beauty"), ("shop", "cosmetics")],
    },
    "Cleaning Companies": {
        "yell": ["cleaning company", "domestic cleaner", "office cleaning", "carpet cleaning"],
        "osm":  [],
    },
    "Locksmiths": {
        "yell": ["locksmith", "emergency locksmith", "lock repairs"],
        "osm":  [("craft", "locksmith")],
    },
    "Car Detailers": {
        "yell": ["car detailing", "car valeting", "mobile valeting"],
        "osm":  [("shop", "car_wash")],
    },
    "Dog Groomers": {
        "yell": ["dog groomer", "dog grooming", "pet grooming"],
        "osm":  [("shop", "pet_grooming")],
    },
}

SOCIAL_DOMAINS = ("facebook.com", "instagram.com", "twitter.com", "tiktok.com", "linkedin.com", "linktree.com")
BROWSER_HDR = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"}
YELL_HDR = {**BROWSER_HDR, "Accept": "text/html,application/xhtml+xml;q=0.9,*/*;q=0.8", "Accept-Language": "en-GB,en;q=0.9", "Referer": "https://www.yell.com/"}
OUTDATED_YEARS = 3
CONTACT_PATHS = ["/contact", "/contact-us", "/contact.html", "/about", "/about-us", "/get-in-touch"]
EMAIL_RE = re.compile(r'\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b')
SKIP_EM = {"noreply", "no-reply", "example", "test", "wordpress", "sentry", "privacy", "abuse", "postmaster", "webmaster", "donotreply"}
_PC_RE = re.compile(r'\b([A-Z]{1,2}\d{1,2}[A-Z]?\s?\d[A-Z]{2})\b', re.I)

# ── Helpers ───────────────────────────────────────────────────────────────────

def haversine_mi(lat1, lng1, lat2, lng2):
    R = 3958.8
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp, dl = math.radians(lat2 - lat1), math.radians(lng2 - lng1)
    a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2 * R * math.asin(math.sqrt(a))

def _norm(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]", "", s.lower())

def _postcode(addr):
    m = _PC_RE.search(addr)
    return _norm(m.group(1)) if m else ""

def _osm_address(tags):
    parts = [tags.get("addr:housenumber", ""), tags.get("addr:street", ""),
             tags.get("addr:city", "") or tags.get("addr:town", ""), tags.get("addr:postcode", "")]
    return ", ".join(p for p in parts if p)

def _get_html(url, timeout=10):
    try:
        r = requests.get(url, headers=BROWSER_HDR, timeout=timeout, allow_redirects=True)
        if r.status_code == 200:
            return r.text
    except Exception:
        pass
    return ""

def _extract_email(html):
    if not html:
        return ""
    soup = BeautifulSoup(html, "lxml")
    for tag in soup.find_all("a", href=re.compile(r"^mailto:", re.I)):
        m = re.search(r"mailto:([^?&\s]+)", tag["href"])
        if m:
            c = m.group(1).strip().lower()
            if not any(s in c for s in SKIP_EM) and "@" in c:
                return c
    for c in EMAIL_RE.findall(html):
        if not any(s in c.lower() for s in SKIP_EM) and "." in c.split("@")[-1]:
            return c.lower()
    return ""

# ── Data fetching ─────────────────────────────────────────────────────────────

def fetch_from_overpass(osm_tags, lat, lng, radius_m):
    if not osm_tags:
        return []
    tag_lines = ""
    for k, v in osm_tags:
        tag_lines += f'node["{k}"="{v}"](around:{radius_m},{lat},{lng});\n'
        tag_lines += f'way["{k}"="{v}"](around:{radius_m},{lat},{lng});\n'
    query = f"[out:json][timeout:90];\n(\n{tag_lines});\nout center tags;"
    try:
        resp = requests.post("https://overpass-api.de/api/interpreter",
                             data={"data": query}, headers={"User-Agent": "UKLeadFinder/2.0"}, timeout=120)
        resp.raise_for_status()
    except Exception as e:
        print(f"    OSM error: {e}")
        return []
    results = []
    radius_mi = radius_m / 1609.344
    for el in resp.json().get("elements", []):
        tags = el.get("tags", {})
        name = tags.get("name", "").strip()
        if not name:
            continue
        if el["type"] == "node":
            blat, blng = el.get("lat"), el.get("lon")
        else:
            c = el.get("center", {})
            blat, blng = c.get("lat"), c.get("lon")
        if blat is None:
            continue
        dist = haversine_mi(lat, lng, blat, blng)
        if dist > radius_mi:
            continue
        results.append({
            "name": name,
            "phone": tags.get("phone") or tags.get("contact:phone", ""),
            "website": tags.get("website") or tags.get("contact:website", ""),
            "address": _osm_address(tags),
            "distance_mi": round(dist, 1),
            "source": "osm",
            "yell_url": "",
        })
    return results

def fetch_from_yell(keywords, city_name, radius_mi):
    all_results = []
    for keyword in keywords:
        print(f"    Yell: '{keyword}' ", end="", flush=True)
        for page_num in range(1, 16):
            try:
                resp = requests.get(
                    "https://www.yell.com/ucs/UcsSearchAction.do",
                    params={"keywords": keyword, "location": city_name, "radius": str(int(radius_mi)), "pageNum": str(page_num)},
                    headers=YELL_HDR, timeout=20)
                if resp.status_code != 200:
                    break
                html = resp.text
            except Exception:
                break
            soup = BeautifulSoup(html, "lxml")
            arts = soup.find_all("article", class_=lambda c: c and "businessCapsule--listing" in c)
            if not arts:
                break
            for art in arts:
                def _t(cls):
                    tag = art.find(class_=lambda c: c and cls in c)
                    return tag.get_text(strip=True) if tag else ""
                name = _t("businessCapsule--name")
                if not name:
                    continue
                site_tag = art.find("a", class_=lambda c: c and "businessCapsule--website" in c)
                website = (site_tag.get("data-url") or site_tag.get("href", "")) if site_tag else ""
                name_tag = art.find(class_=lambda c: c and "businessCapsule--name" in c)
                yell_url = ""
                if name_tag:
                    a_tag = name_tag if name_tag.name == "a" else name_tag.find("a")
                    if a_tag and a_tag.get("href"):
                        yell_url = urljoin("https://www.yell.com", a_tag["href"])
                dist_txt = _t("businessCapsule--distance")
                dist_mi = 0.0
                m = re.search(r"([\d.]+)\s*miles?", dist_txt, re.I)
                if m:
                    dist_mi = float(m.group(1))
                if dist_mi <= radius_mi:
                    all_results.append({
                        "name": name,
                        "phone": _t("businessCapsule--telephone"),
                        "website": website,
                        "address": _t("businessCapsule--address"),
                        "distance_mi": dist_mi,
                        "source": "yell",
                        "yell_url": yell_url,
                    })
            print(".", end="", flush=True)
            if len(arts) < 10:
                break
            time.sleep(1.5)
        print()
    return all_results

def deduplicate(items):
    seen, merged = {}, []
    for biz in items:
        words = biz["name"].split()
        first = _norm(words[0]) if words else ""
        pc = _postcode(biz["address"])
        key = (first, pc) if pc else (_norm(biz["name"]),)
        if key not in seen:
            seen[key] = True
            merged.append(biz)
    return merged

def check_website(url):
    if not url:
        return "none", "No website listed"
    if any(d in url.lower() for d in SOCIAL_DOMAINS):
        d = next(x for x in SOCIAL_DOMAINS if x in url.lower())
        return "social_only", f"Only a {d} page"
    try:
        resp = requests.get(url, headers=BROWSER_HDR, timeout=12, allow_redirects=True)
    except requests.exceptions.SSLError:
        return "outdated", "Broken SSL"
    except Exception:
        return "error", "Cannot connect"
    if resp.status_code >= 400:
        return "error", f"HTTP {resp.status_code}"
    cur = datetime.now().year
    text = resp.text
    yrs = re.findall(r'copyright[^\d]{0,10}(\d{4})', text, re.I) + re.findall(r'[©]\s*(\d{4})', text)
    valid = [int(y) for y in yrs if 2000 <= int(y) <= cur + 1]
    if valid and cur - max(valid) >= OUTDATED_YEARS:
        return "outdated", f"Copyright: {max(valid)}"
    return "active", "Website looks current"

def scrape_contacts(website_url, yell_url=""):
    email = ""
    if website_url and not any(d in website_url.lower() for d in SOCIAL_DOMAINS):
        pages = [website_url] + [urljoin(website_url, p) for p in CONTACT_PATHS]
        for page_url in pages:
            if email:
                break
            email = _extract_email(_get_html(page_url))
    if not email and yell_url:
        email = _extract_email(_get_html(yell_url))
    return email

# ── Excel output ──────────────────────────────────────────────────────────────

def save_leads(leads, city_name, output_dir):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = output_dir / f"leads_{_norm(city_name)}_{ts}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    FILL_H = PatternFill("solid", fgColor="1A2035")
    FILLS = {
        "none":        PatternFill("solid", fgColor="FFD6D6"),
        "social_only": PatternFill("solid", fgColor="D6EAFF"),
        "outdated":    PatternFill("solid", fgColor="FFF2CC"),
        "error":       PatternFill("solid", fgColor="E8E8E8"),
    }
    THIN = Border(**{s: Side(style="thin", color="CCCCCC") for s in ("left", "right", "top", "bottom")})

    COLS = [
        ("Business Name", 32), ("Industry", 24), ("City", 16),
        ("Phone", 18), ("Email", 34), ("Website Status", 18),
        ("Website / Social URL", 40), ("Address", 42),
        ("Notes", 30), ("Distance (mi)", 14), ("Yell Listing", 38),
    ]

    for col, (hdr, w) in enumerate(COLS, 1):
        c = ws.cell(1, col, hdr)
        c.fill = FILL_H
        c.font = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28

    ORDER = {"none": 0, "social_only": 1, "outdated": 2, "error": 3}
    sorted_leads = sorted(leads, key=lambda b: (ORDER.get(b["website_status"], 9), b.get("distance_mi", 99)))

    for ri, biz in enumerate(sorted_leads, 2):
        fill = FILLS.get(biz["website_status"], PatternFill(fill_type=None))
        row = [
            biz.get("name", ""),
            biz.get("business_type", ""),
            city_name.title(),
            biz.get("phone", ""),
            biz.get("email", ""),
            biz.get("website_status", "").upper().replace("_", " "),
            biz.get("website", ""),
            biz.get("address", ""),
            biz.get("status_notes", ""),
            round(float(biz.get("distance_mi", 0)), 1),
            biz.get("yell_url", ""),
        ]
        for col, val in enumerate(row, 1):
            c = ws.cell(ri, col, val)
            c.fill = fill
            c.alignment = Alignment(vertical="center")
            c.border = THIN
        ws.row_dimensions[ri].height = 17

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    cnts = {s: sum(1 for b in sorted_leads if b["website_status"] == s) for s in ("none", "social_only", "outdated", "error")}
    summary_rows = [
        ("City", city_name.title()),
        ("Date", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Sources", "OpenStreetMap + Yell.com"),
        ("", ""),
        ("Total leads", len(sorted_leads)),
        ("No website (RED)", cnts["none"]),
        ("Social only (BLUE)", cnts["social_only"]),
        ("Outdated (AMBER)", cnts["outdated"]),
        ("With email", sum(1 for b in sorted_leads if b.get("email"))),
        ("With phone", sum(1 for b in sorted_leads if b.get("phone"))),
    ]
    for r, (label, val) in enumerate(summary_rows, 1):
        ws2.cell(r, 1, label).font = Font(bold=True, size=10)
        ws2.cell(r, 2, val)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 36

    wb.save(filename)
    return filename, sorted_leads

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Find UK leads with no/weak website")
    parser.add_argument("--city", default="wigan", help="UK city name (default: wigan)")
    parser.add_argument("--radius", type=int, default=10, help="Search radius in miles (default: 10)")
    parser.add_argument("--types", default="all", help='Comma-separated business types or "all"')
    args = parser.parse_args()

    city_key = args.city.lower().strip()
    if city_key in UK_CITIES:
        lat, lng = UK_CITIES[city_key]
    else:
        print(f"  City '{args.city}' not in preset list. Available cities:")
        print("  " + ", ".join(sorted(UK_CITIES.keys())))
        print("\n  Defaulting to Wigan.")
        lat, lng = UK_CITIES["wigan"]
        city_key = "wigan"

    radius_mi = args.radius
    radius_m = int(radius_mi * 1609.344)

    if args.types.lower() == "all":
        selected = PRESETS
    else:
        keys = [t.strip() for t in args.types.split(",")]
        selected = {k: v for k, v in PRESETS.items() if any(k.lower().startswith(t.lower()) for t in keys)}
        if not selected:
            selected = PRESETS

    output_dir = Path(__file__).parent
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*60}")
    print(f"  L&D Designs — UK Lead Finder")
    print(f"  City: {city_key.title()}  |  Radius: {radius_mi} miles  |  Types: {len(selected)}")
    print(f"{'='*60}\n")

    global_seen = set()
    all_raw = []

    for type_label, cfg in selected.items():
        print(f"\n  [{type_label}]")
        osm = fetch_from_overpass(cfg["osm"], lat, lng, radius_m)
        yell = fetch_from_yell(cfg["yell"], f"{city_key.title()}, UK", radius_mi)
        combined = deduplicate(osm + yell)
        added = 0
        for biz in combined:
            words = biz["name"].split()
            first = _norm(words[0]) if words else ""
            pc = _postcode(biz["address"])
            key = (first, pc) if pc else (_norm(biz["name"]),)
            if key not in global_seen:
                global_seen.add(key)
                biz["business_type"] = type_label
                all_raw.append(biz)
                added += 1
        print(f"    +{added} new  |  Total so far: {len(all_raw)}")

    print(f"\n  Checking websites and finding emails for {len(all_raw)} businesses...\n")
    leads = []
    for idx, biz in enumerate(all_raw, 1):
        status, notes = check_website(biz.get("website", ""))
        if idx % 25 == 0:
            print(f"  Checked {idx}/{len(all_raw)}...")
        if status == "active":
            continue
        email = scrape_contacts(biz.get("website", ""), biz.get("yell_url", ""))
        leads.append({**biz, "website_status": status, "status_notes": notes, "email": email})

    print(f"\n  Saving {len(leads)} leads...")
    filename, sorted_leads = save_leads(leads, city_key, output_dir)

    cnts = {s: sum(1 for b in sorted_leads if b["website_status"] == s) for s in ("none", "social_only", "outdated", "error")}

    print(f"\n{'='*60}")
    print(f"  Done! Saved: {filename.name}")
    print(f"  No website   : {cnts['none']}")
    print(f"  Social only  : {cnts['social_only']}")
    print(f"  Outdated     : {cnts['outdated']}")
    print(f"  With email   : {sum(1 for b in sorted_leads if b.get('email'))}")
    print(f"  With phone   : {sum(1 for b in sorted_leads if b.get('phone'))}")
    print(f"{'='*60}\n")

    # Update progress tracker
    progress_file = Path(__file__).parent.parent / "progress.json"
    if progress_file.exists():
        try:
            progress = json.loads(progress_file.read_text())
        except Exception:
            progress = {}
    else:
        progress = {}

    progress.setdefault("lead_searches", [])
    progress["lead_searches"].append({
        "date": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "city": city_key.title(),
        "radius_mi": radius_mi,
        "total_found": len(sorted_leads),
        "no_website": cnts["none"],
        "social_only": cnts["social_only"],
        "with_email": sum(1 for b in sorted_leads if b.get("email")),
        "file": filename.name,
    })
    progress["last_updated"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    progress_file.write_text(json.dumps(progress, indent=2))

    return filename

if __name__ == "__main__":
    main()
