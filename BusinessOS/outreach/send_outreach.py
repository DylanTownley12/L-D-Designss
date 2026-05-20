#!/usr/bin/env python3
"""
Outreach Sender — L&D Designs
Reads from master CRM and sends emails/WhatsApp messages safely.

Rules:
- Max 30 emails/day (Gmail safe limit for cold outreach)
- Never send duplicate messages
- Logs every action to CRM
- 3-second delay between sends

Usage:
    python send_outreach.py --dry-run           # Preview without sending
    python send_outreach.py --limit 10          # Send to 10 leads
    python send_outreach.py --channel email     # Email only
    python send_outreach.py --channel whatsapp  # WhatsApp links only (manual)
"""

import subprocess, sys, argparse, json
from pathlib import Path
from datetime import datetime

def _ensure(pkg, import_as=None):
    try:
        __import__(import_as or pkg)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", pkg])

_ensure("openpyxl")

import openpyxl
from openpyxl.styles import PatternFill

sys.path.insert(0, str(Path(__file__).parent))
from templates import get_email, get_whatsapp, choose_template

CRM_FILE     = Path(__file__).parent.parent / "leads" / "master_crm.xlsx"
LOG_FILE     = Path(__file__).parent / "outreach_log.json"
MAX_PER_DAY  = 30
DAILY_COUNTER_FILE = Path(__file__).parent / ".daily_count.json"

STATUS_COL   = 20  # "Outreach Status" column index (1-based)
EMAIL_COL    = 5
PHONE_COL    = 6
NAME_COL     = 1
INDUSTRY_COL = 2
CONTACT_COL  = 11  # First Contact Sent
LAST_COL     = 18  # Last Contact Date
NEXT_COL     = 19  # Next Follow-up Date


def load_log():
    if LOG_FILE.exists():
        try:
            return json.loads(LOG_FILE.read_text())
        except Exception:
            pass
    return {"sent": [], "total_sent": 0}

def save_log(log):
    LOG_FILE.write_text(json.dumps(log, indent=2))

def get_daily_count():
    today = datetime.now().strftime("%Y-%m-%d")
    if DAILY_COUNTER_FILE.exists():
        try:
            data = json.loads(DAILY_COUNTER_FILE.read_text())
            if data.get("date") == today:
                return data.get("count", 0)
        except Exception:
            pass
    return 0

def increment_daily_count():
    today = datetime.now().strftime("%Y-%m-%d")
    count = get_daily_count() + 1
    DAILY_COUNTER_FILE.write_text(json.dumps({"date": today, "count": count}))
    return count

def already_contacted(log, business_name):
    return any(e["business"] == business_name for e in log.get("sent", []))

def load_crm():
    if not CRM_FILE.exists():
        print(f"  CRM not found at {CRM_FILE}")
        print("  Run: python BusinessOS/leads/create_crm.py")
        return None, None
    wb = openpyxl.load_workbook(CRM_FILE)
    ws = wb["CRM"]
    return wb, ws

def get_leads_to_contact(ws, limit):
    leads = []
    for row in ws.iter_rows(min_row=2):
        name   = row[NAME_COL - 1].value
        status = row[STATUS_COL - 1].value
        email  = row[EMAIL_COL - 1].value
        if not name:
            continue
        if status in (None, "", "Not Contacted"):
            leads.append({
                "row":      row[0].row,
                "name":     str(name),
                "email":    str(email) if email else "",
                "phone":    str(row[PHONE_COL - 1].value or ""),
                "industry": str(row[INDUSTRY_COL - 1].value or ""),
                "website_status": str(row[6].value or "none").lower().replace(" ", "_"),
            })
        if len(leads) >= limit:
            break
    return leads

def mark_contacted(ws, row_num):
    ws.cell(row_num, STATUS_COL).value  = "Contacted"
    ws.cell(row_num, CONTACT_COL).value = datetime.now().strftime("%d/%m/%Y")
    ws.cell(row_num, LAST_COL).value    = datetime.now().strftime("%d/%m/%Y")
    from datetime import timedelta
    next_date = datetime.now() + timedelta(days=3)
    ws.cell(row_num, NEXT_COL).value    = next_date.strftime("%d/%m/%Y")
    ws.cell(row_num, STATUS_COL).fill   = PatternFill("solid", fgColor="D6EAFF")

def preview_outreach(leads):
    print(f"\n  {'='*56}")
    print(f"  DRY RUN — {len(leads)} leads queued")
    print(f"  {'='*56}")
    for i, lead in enumerate(leads, 1):
        tkey = choose_template(lead["website_status"])
        e = get_email(tkey, lead["name"])
        print(f"\n  [{i}] {lead['name']}")
        print(f"      Email: {lead['email'] or 'none'}")
        print(f"      Template: {tkey}")
        print(f"      Subject: {e['subject']}")
    print(f"\n  No messages sent (dry-run mode).\n")

def send_email_gmail(to_email, subject, body):
    """
    Send via Gmail SMTP.
    Set GMAIL_ADDRESS and GMAIL_APP_PASSWORD environment variables.
    Generate app password at: myaccount.google.com/apppasswords
    """
    import smtplib, os
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    gmail_addr = os.getenv("GMAIL_ADDRESS")
    gmail_pass = os.getenv("GMAIL_APP_PASSWORD")

    if not gmail_addr or not gmail_pass:
        raise EnvironmentError(
            "Set GMAIL_ADDRESS and GMAIL_APP_PASSWORD environment variables.\n"
            "Get app password: myaccount.google.com/apppasswords"
        )

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = gmail_addr
    msg["To"]      = to_email
    msg.attach(MIMEText(body, "plain"))

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(gmail_addr, gmail_pass)
        server.sendmail(gmail_addr, to_email, msg.as_string())

def run(dry_run, limit, channel):
    log = load_log()
    daily_count = get_daily_count()

    wb, ws = load_crm()
    if ws is None:
        return

    remaining_today = MAX_PER_DAY - daily_count
    if remaining_today <= 0:
        print(f"  Daily limit ({MAX_PER_DAY}) reached. Try again tomorrow.")
        return

    actual_limit = min(limit, remaining_today)
    leads = get_leads_to_contact(ws, actual_limit * 3)  # fetch extra in case some have no email

    if channel == "email":
        leads = [l for l in leads if l["email"]][:actual_limit]

    if not leads:
        print("  No leads to contact. Run the lead finder first.")
        return

    if dry_run:
        preview_outreach(leads)
        return

    sent_count = 0
    for lead in leads:
        if sent_count >= actual_limit:
            break
        if already_contacted(log, lead["name"]):
            continue

        tkey = choose_template(lead["website_status"])

        try:
            if channel == "email" and lead["email"]:
                e = get_email(tkey, lead["name"])
                send_email_gmail(lead["email"], e["subject"], e["body"])
                mark_contacted(ws, lead["row"])
                log["sent"].append({
                    "business": lead["name"],
                    "channel":  "email",
                    "to":       lead["email"],
                    "template": tkey,
                    "date":     datetime.now().strftime("%d/%m/%Y %H:%M"),
                })
                sent_count += 1
                increment_daily_count()
                print(f"  Sent email → {lead['name']} ({lead['email']})")

            elif channel == "whatsapp":
                msg = get_whatsapp(tkey, lead["name"])
                mark_contacted(ws, lead["row"])
                log["sent"].append({
                    "business": lead["name"],
                    "channel":  "whatsapp",
                    "to":       lead["phone"],
                    "template": tkey,
                    "date":     datetime.now().strftime("%d/%m/%Y %H:%M"),
                    "message":  msg,
                })
                sent_count += 1
                print(f"  Logged WhatsApp → {lead['name']} ({lead['phone']})")
                print(f"  Message: {msg[:80]}...")

        except Exception as e:
            print(f"  Failed → {lead['name']}: {e}")
            continue

        import time
        time.sleep(3)

    log["total_sent"] = log.get("total_sent", 0) + sent_count
    log["last_run"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    save_log(log)
    wb.save(CRM_FILE)

    print(f"\n  Done. Sent {sent_count} messages today (daily total: {daily_count + sent_count}/{MAX_PER_DAY})")

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run",  action="store_true", help="Preview only, no messages sent")
    parser.add_argument("--limit",    type=int, default=10, help="Max leads to contact (default 10)")
    parser.add_argument("--channel",  choices=["email", "whatsapp"], default="email")
    args = parser.parse_args()

    print(f"\n  L&D Designs — Outreach Sender")
    print(f"  Channel: {args.channel}  |  Limit: {args.limit}  |  Dry run: {args.dry_run}\n")
    run(args.dry_run, args.limit, args.channel)

if __name__ == "__main__":
    main()
