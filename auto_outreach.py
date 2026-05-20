#!/usr/bin/env python3
"""
L&D Designs — Automated Email & SMS Outreach
Reads credentials from .env, finds the latest leads spreadsheet, sends.

Usage:
    python auto_outreach.py                        # uses latest wigan_hair_leads_*.xlsx
    python auto_outreach.py leads_file.xlsx        # explicit spreadsheet
"""

import subprocess, sys

def _ensure(package, import_as=None):
    try:
        __import__(import_as or package)
    except ImportError:
        print(f"  Installing {package} …")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", package])

_ensure("openpyxl")
_ensure("twilio")
_ensure("dotenv", "dotenv")

import os, re, csv, glob, smtplib, time, random
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

import openpyxl

# ── Credentials ───────────────────────────────────────────────────────────────

GMAIL_ADDRESS      = os.getenv("GMAIL_ADDRESS", "")
GMAIL_APP_PASSWORD = os.getenv("GMAIL_APP_PASSWORD", "")
TWILIO_ACCOUNT_SID = os.getenv("TWILIO_ACCOUNT_SID", "")
TWILIO_AUTH_TOKEN  = os.getenv("TWILIO_AUTH_TOKEN", "")
TWILIO_FROM_NUMBER = os.getenv("TWILIO_FROM_NUMBER", "")
YOUR_NAME          = os.getenv("YOUR_NAME", "Dylan")
YOUR_PHONE         = os.getenv("YOUR_PHONE", "07504683058")
SHOWCASE_LINK      = os.getenv("SHOWCASE_LINK", "https://mellow-speculoos-d1850c.netlify.app/showcase.html")
MAX_EMAILS         = int(os.getenv("MAX_EMAILS", "30"))
MAX_SMS            = int(os.getenv("MAX_SMS", "20"))

USE_EMAIL = bool(GMAIL_ADDRESS and GMAIL_APP_PASSWORD
                 and GMAIL_ADDRESS != "your@gmail.com"
                 and GMAIL_APP_PASSWORD != "xxxx xxxx xxxx xxxx")
USE_SMS   = bool(TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN and TWILIO_FROM_NUMBER)

# ── Message templates ─────────────────────────────────────────────────────────

EMAIL_SUBJECT = "Quick question — website for {name}"

EMAIL_BODY = """\
Hi,

I noticed {name} doesn't have a website yet.

I build professional websites for local businesses in Wigan from just £199 — usually done within a week. No monthly fees, one-off payment.

Here's some of my work: {showcase}

Would you be interested in a free quote?

{your_name}
L&D Designs
WhatsApp / Call: {your_phone}"""

SMS_BODY = ("Hi, I noticed {name} doesn't have a website. "
            "I build professional sites for Wigan businesses from £199 — "
            "done in days, no monthly fees. Free quote? "
            "- {your_name}, L&D Designs")

# ── Helpers ───────────────────────────────────────────────────────────────────

def gf(lead, *keys):
    for k in keys:
        v = lead.get(k)
        if v and str(v).strip().lower() not in ("", "none", "nan", "null"):
            return str(v).strip()
    return ""


def clean_phone(p):
    d = re.sub(r"[^\d+]", "", p)
    if d.startswith("0"):
        d = "+44" + d[1:]
    elif d and not d.startswith("+"):
        d = "+44" + d
    return d


def find_spreadsheet():
    files = sorted(glob.glob("wigan_hair_leads_*.xlsx"), reverse=True)
    if files:
        return files[0]
    files = sorted(glob.glob("multi_leads_*.xlsx"), reverse=True)
    if files:
        return files[0]
    return None


def load_leads(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            rows.append(dict(zip(headers, row)))
    return rows

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print()
    print("╔══════════════════════════════════════════════════════════════╗")
    print("║   L&D Designs — Automated Outreach                          ║")
    print("╚══════════════════════════════════════════════════════════════╝")
    print()

    # Config check
    print(f"  Email : {'✅ Configured' if USE_EMAIL else '❌ Not configured — fill in .env'}")
    print(f"  SMS   : {'✅ Configured' if USE_SMS else '⚪ Not configured (optional)'}")
    print()

    if not USE_EMAIL:
        print("  Copy .env.example to .env and fill in your Gmail details.")
        sys.exit(1)

    # Find spreadsheet
    xlsx = sys.argv[1] if len(sys.argv) > 1 else find_spreadsheet()
    if not xlsx or not Path(xlsx).exists():
        print("  No leads spreadsheet found. Run find_leads.py first.")
        sys.exit(1)

    print(f"  Loading: {xlsx}")
    all_leads = load_leads(xlsx)
    print(f"  Rows: {len(all_leads)}")

    # Filter
    email_leads, sms_leads = [], []
    for lead in all_leads:
        status = gf(lead, "Website Status", "website_status").upper()
        if "ACTIVE" in status:
            continue
        email = gf(lead, "Email Address", "Email", "email")
        phone = gf(lead, "Phone Number", "Phone", "phone")
        if email and "@" in email:
            email_leads.append(lead)
        elif phone:
            sms_leads.append(lead)

    print()
    print(f"  Ready to email : {len(email_leads)}")
    print(f"  Ready to SMS   : {len(sms_leads)}  (no email, has phone)")
    skipped = len(all_leads) - len(email_leads) - len(sms_leads)
    print(f"  Skipped        : {skipped}  (has active website or no contact info)")
    print(f"  Will send      : up to {MAX_EMAILS} emails + {MAX_SMS} SMS")
    print()

    log = []
    sent_email = sent_sms = failed = 0

    # ── Email ─────────────────────────────────────────────────────────────────
    batch = email_leads[:MAX_EMAILS]
    if batch:
        print(f"Connecting to Gmail SMTP …", end=" ", flush=True)
        try:
            server = smtplib.SMTP_SSL("smtp.gmail.com", 465)
            server.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
            print("✅ Connected.")
        except Exception as e:
            print(f"\n❌ Gmail login failed: {e}")
            print("  Check GMAIL_ADDRESS and GMAIL_APP_PASSWORD in .env")
            sys.exit(1)

        print(f"\nSending {len(batch)} emails …\n")
        for i, lead in enumerate(batch):
            name  = gf(lead, "Business Name", "name") or "there"
            email = gf(lead, "Email Address", "Email", "email")
            ts    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            subject = EMAIL_SUBJECT.format(name=name)
            body    = EMAIL_BODY.format(name=name, showcase=SHOWCASE_LINK,
                                        your_name=YOUR_NAME, your_phone=YOUR_PHONE)
            msg = MIMEMultipart()
            msg["From"]    = GMAIL_ADDRESS
            msg["To"]      = email
            msg["Subject"] = subject
            msg.attach(MIMEText(body, "plain"))

            try:
                server.sendmail(GMAIL_ADDRESS, email, msg.as_string())
                sent_email += 1
                status = "sent"
                print(f"  [{i+1}/{len(batch)}] ✅  {name} → {email}")
            except Exception as e:
                failed += 1
                status = f"failed: {e}"
                print(f"  [{i+1}/{len(batch)}] ❌  {name} → {email} — {e}")

            log.append({"type": "email", "name": name, "contact": email,
                        "status": status, "timestamp": ts})

            if i < len(batch) - 1 and status == "sent":
                time.sleep(random.uniform(3, 6))

        server.quit()
        print(f"\n  Email batch done. Sent: {sent_email}  Failed: {failed}")

    # ── SMS ───────────────────────────────────────────────────────────────────
    if USE_SMS and sms_leads:
        from twilio.rest import Client as TwilioClient
        tw = TwilioClient(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
        sms_batch = sms_leads[:MAX_SMS]
        print(f"\nSending {len(sms_batch)} SMS messages …\n")

        for i, lead in enumerate(sms_batch):
            name  = gf(lead, "Business Name", "name") or "there"
            phone = clean_phone(gf(lead, "Phone Number", "Phone", "phone"))
            ts    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            body  = SMS_BODY.format(name=name, your_name=YOUR_NAME)

            try:
                tw.messages.create(body=body, from_=TWILIO_FROM_NUMBER, to=phone)
                sent_sms += 1
                status = "sent"
                print(f"  [{i+1}/{len(sms_batch)}] ✅  {name} → {phone}")
            except Exception as e:
                failed += 1
                status = f"failed: {e}"
                print(f"  [{i+1}/{len(sms_batch)}] ❌  {name} → {phone} — {e}")

            log.append({"type": "sms", "name": name, "contact": phone,
                        "status": status, "timestamp": ts})
            time.sleep(1)

        print(f"\n  SMS batch done. Sent: {sent_sms}")

    # ── Save log ──────────────────────────────────────────────────────────────
    log_file = f"outreach_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    with open(log_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["type", "name", "contact", "status", "timestamp"])
        writer.writeheader()
        writer.writerows(log)

    success = sum(1 for r in log if r["status"] == "sent")
    print()
    print("╔══════════════════════════════════════════════════════════════╗")
    print(f"║  TOTAL: {sent_email} emails + {sent_sms} SMS sent                            ║")
    print(f"║  Success rate: {success}/{len(log)}                                       ║")
    print(f"║  Log saved: {log_file:<49}║")
    print("╚══════════════════════════════════════════════════════════════╝")
    print()


if __name__ == "__main__":
    main()
