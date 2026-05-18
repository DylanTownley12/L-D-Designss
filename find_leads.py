#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════╗
║   Barber & Hairdresser Lead Finder – 50 miles of Wigan       ║
║   No API key needed – uses OpenStreetMap + Yell.com           ║
╠══════════════════════════════════════════════════════════════╣
║  HOW TO RUN:                                                 ║
║    Windows  → double-click  run.bat                          ║
║    Mac/Linux → double-click run.sh  or  python3 find_leads.py║
║                                                              ║
║  The script installs any missing packages automatically.     ║
╚══════════════════════════════════════════════════════════════╝
"""

# =========================================================================
# Auto-install missing packages — user never needs to open a terminal
# =========================================================================
import subprocess, sys

def _ensure(package, import_as=None):
    try:
        __import__(import_as or package)
    except ImportError:
        print(f"  Installing {package} …")
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "-q", package],
            stdout=subprocess.DEVNULL,
        )

_ensure("requests")
_ensure("openpyxl")
_ensure("beautifulsoup4", "bs4")
_ensure("lxml")

# =========================================================================
# Imports
# =========================================================================
import re
import math
import time
import unicodedata
from datetime import datetime
from urllib.parse import urljoin

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

# =========================================================================
# Configuration
# =========================================================================

WIGAN_LAT  = 53.5450
WIGAN_LNG  = -2.6325
RADIUS_MI  = 50
RADIUS_M   = int(RADIUS_MI * 1609.344)       # 80 467 m

# ── Overpass API (OpenStreetMap) ──────────────────────────────────────────
OVERPASS_URL = "https://overpass-api.de/api/interpreter"

# ── Yell.com ──────────────────────────────────────────────────────────────
YELL_BASE      = "https://www.yell.com/ucs/UcsSearchAction.do"
YELL_KEYWORDS  = ["barbers", "hairdressers", "hair salon", "barbershop"]
YELL_MAX_PAGES = 20   # ~10 listings/page → up to 200 results per keyword

YELL_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-GB,en;q=0.9",
    "Referer": "https://www.yell.com/",
}

# Website is "outdated" if copyright / Last-Modified year is this far behind
OUTDATED_YEARS = 3

BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}

SOCIAL_DOMAINS = (
    "facebook.com", "instagram.com", "twitter.com",
    "tiktok.com", "linkedin.com", "linktree.com",
)

# =========================================================================
# Geo helper
# =========================================================================

def haversine_mi(lat1, lng1, lat2, lng2):
    R = 3_958.8
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lng2 - lng1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))

# =========================================================================
# Source 1 — Overpass API (OpenStreetMap)
# =========================================================================

def _osm_address(tags):
    parts = [
        tags.get("addr:housenumber", ""),
        tags.get("addr:street", ""),
        tags.get("addr:city", "") or tags.get("addr:town", ""),
        tags.get("addr:postcode", ""),
    ]
    return ", ".join(p for p in parts if p)


def fetch_from_overpass():
    """Single POST request — returns all OSM barbers/hairdressers within 50 mi of Wigan."""
    query = f"""
[out:json][timeout:90];
(
  node["shop"="hairdresser"](around:{RADIUS_M},{WIGAN_LAT},{WIGAN_LNG});
  way["shop"="hairdresser"](around:{RADIUS_M},{WIGAN_LAT},{WIGAN_LNG});
  node["shop"="barber"](around:{RADIUS_M},{WIGAN_LAT},{WIGAN_LNG});
  way["shop"="barber"](around:{RADIUS_M},{WIGAN_LAT},{WIGAN_LNG});
  node["amenity"="hairdresser"](around:{RADIUS_M},{WIGAN_LAT},{WIGAN_LNG});
  way["amenity"="hairdresser"](around:{RADIUS_M},{WIGAN_LAT},{WIGAN_LNG});
);
out center tags;
""".strip()

    print("  Querying OpenStreetMap (Overpass API) …", end="", flush=True)
    try:
        resp = requests.post(
            OVERPASS_URL,
            data={"data": query},
            headers={"User-Agent": "WiganLeadFinder/1.0"},
            timeout=120,
        )
        resp.raise_for_status()
    except Exception as exc:
        print(f"\n  [!] Overpass error: {exc}")
        return []

    elements = resp.json().get("elements", [])
    print(f" {len(elements)} raw elements returned.")

    results = []
    for el in elements:
        tags = el.get("tags", {})
        name = tags.get("name", "").strip()
        if not name:
            continue

        # node → direct lat/lon; way → center lat/lon (from `out center tags`)
        if el["type"] == "node":
            blat = el.get("lat")
            blng = el.get("lon")
        else:
            center = el.get("center", {})
            blat = center.get("lat")
            blng = center.get("lon")

        if blat is None or blng is None:
            continue

        dist = haversine_mi(WIGAN_LAT, WIGAN_LNG, blat, blng)
        if dist > RADIUS_MI:
            continue   # secondary check — Overpass `around:` is approximate for ways

        phone   = tags.get("phone") or tags.get("contact:phone", "")
        website = tags.get("website") or tags.get("contact:website", "")

        results.append({
            "name":        name,
            "phone":       phone,
            "website":     website,
            "address":     _osm_address(tags),
            "rating":      "",
            "reviews":     0,
            "distance_mi": round(dist, 1),
            "lat":         blat,
            "lng":         blng,
            "source":      "osm",
        })

    print(f"  Overpass: {len(results)} named businesses within {RADIUS_MI} miles.")
    return results

# =========================================================================
# Source 2 — Yell.com scraping
# =========================================================================

def _fetch_yell_page(keyword, page_num):
    params = {
        "keywords": keyword,
        "location": "Wigan, Greater Manchester",
        "radius":   "50",
        "pageNum":  str(page_num),
    }
    try:
        resp = requests.get(
            YELL_BASE, params=params, headers=YELL_HEADERS, timeout=20
        )
        if resp.status_code == 200:
            return resp.text
    except Exception:
        pass
    return ""


def _parse_yell_page(html):
    soup = BeautifulSoup(html, "lxml")

    # Substring class match — resilient to Yell adding modifier classes
    articles = soup.find_all(
        "article",
        class_=lambda c: c and "businessCapsule--listing" in c,
    )

    results = []
    for art in articles:
        def _text(cls_part):
            tag = art.find(class_=lambda c: c and cls_part in c)
            return tag.get_text(strip=True) if tag else ""

        name = _text("businessCapsule--name")
        if not name:
            continue

        address  = _text("businessCapsule--address")
        phone    = _text("businessCapsule--telephone")
        dist_txt = _text("businessCapsule--distance")

        # Website — prefer data-url attribute, fall back to href
        site_tag = art.find("a", class_=lambda c: c and "businessCapsule--website" in c)
        website  = ""
        if site_tag:
            website = site_tag.get("data-url") or site_tag.get("href", "")

        # Parse distance
        dist_mi = 0.0
        m = re.search(r"([\d.]+)\s*miles?", dist_txt, re.I)
        if m:
            dist_mi = float(m.group(1))

        results.append({
            "name":        name,
            "phone":       phone,
            "website":     website,
            "address":     address,
            "rating":      "",
            "reviews":     0,
            "distance_mi": dist_mi,
            "lat":         None,
            "lng":         None,
            "source":      "yell",
        })

    return results


def fetch_from_yell():
    all_results = []

    for keyword in YELL_KEYWORDS:
        print(f"  Yell.com: '{keyword}' ", end="", flush=True)
        keyword_count = 0

        for page_num in range(1, YELL_MAX_PAGES + 1):
            html = _fetch_yell_page(keyword, page_num)
            if not html:
                break

            page_results = _parse_yell_page(html)
            if not page_results:
                break

            filtered = [r for r in page_results if r["distance_mi"] <= RADIUS_MI]
            all_results.extend(filtered)
            keyword_count += len(filtered)
            print(".", end="", flush=True)

            if len(page_results) < 10:   # partial page = last page
                break

            time.sleep(1.5)

        print(f" {keyword_count} listings")

    print(f"  Yell total: {len(all_results)} listings collected.")
    return all_results

# =========================================================================
# Deduplication
# =========================================================================

_UK_PC_RE = re.compile(r'\b([A-Z]{1,2}\d{1,2}[A-Z]?\s?\d[A-Z]{2})\b', re.I)


def _norm(s):
    """Lowercase, ASCII-only, alphanumeric only."""
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]", "", s.lower())


def _postcode(address):
    m = _UK_PC_RE.search(address)
    return _norm(m.group(1)) if m else ""


def deduplicate(osm_list, yell_list):
    """
    Prefer OSM records (richer structured data), use Yell to fill gaps.
    Dedup key: (normalised first word of name, normalised postcode).
    Falls back to full normalised name when postcode is absent from both.
    """
    seen   = {}
    merged = []

    for biz in osm_list + yell_list:
        words = biz["name"].split()
        first = _norm(words[0]) if words else ""
        pc    = _postcode(biz["address"])
        key   = (first, pc) if pc else (_norm(biz["name"]),)

        if key not in seen:
            seen[key] = True
            merged.append(biz)

    return merged

# =========================================================================
# Website status checker
# =========================================================================

def check_website(url):
    """
    Returns (status, notes).
    status: 'none' | 'social_only' | 'outdated' | 'error' | 'active'
    """
    if not url:
        return "none", "No website listed"

    if any(d in url.lower() for d in SOCIAL_DOMAINS):
        domain = next(d for d in SOCIAL_DOMAINS if d in url.lower())
        return "social_only", f"Only a {domain} page"

    try:
        resp = requests.get(url, headers=BROWSER_HEADERS, timeout=12, allow_redirects=True)
    except requests.exceptions.SSLError:
        return "outdated", "Broken SSL certificate"
    except requests.exceptions.ConnectionError:
        return "error", "Cannot connect"
    except requests.exceptions.Timeout:
        return "error", "Timed out"
    except Exception as exc:
        return "error", str(exc)[:80]

    if resp.status_code >= 400:
        return "error", f"HTTP {resp.status_code}"

    current_year = datetime.now().year
    text = resp.text

    # Last-Modified header
    lm = resp.headers.get("Last-Modified", "")
    if lm:
        try:
            from email.utils import parsedate
            parsed = parsedate(lm)
            if parsed and current_year - parsed[0] >= OUTDATED_YEARS:
                return "outdated", f"Last-Modified header says {parsed[0]}"
        except Exception:
            pass

    # Copyright year in page source
    year_hits = (
        re.findall(r'copyright[^\d]{0,10}(\d{4})', text, re.I)
        + re.findall(r'[©]\s*(\d{4})', text)
        + re.findall(r'&copy;\s*(\d{4})', text)
    )
    valid = [int(y) for y in year_hits if 2000 <= int(y) <= current_year + 1]
    if valid and current_year - max(valid) >= OUTDATED_YEARS:
        return "outdated", f"Copyright year: {max(valid)}"

    # Old WordPress via meta generator
    soup = BeautifulSoup(text, "lxml")
    gen  = soup.find("meta", {"name": "generator"})
    if gen:
        content = gen.get("content", "").lower()
        wp = re.search(r"wordpress\s+([\d.]+)", content)
        if wp:
            try:
                if float(wp.group(1)) < 5.0:
                    return "outdated", f"Old WordPress {wp.group(1)}"
            except ValueError:
                pass

    # Table-heavy = old-school layout
    all_tags = soup.find_all()
    if all_tags and len(soup.find_all("table")) / len(all_tags) > 0.12:
        return "outdated", "Table-based layout (very old site)"

    return "active", "Website looks current"

# =========================================================================
# Contact scraper (email + WhatsApp)
# =========================================================================

EMAIL_RE   = re.compile(r'\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b')
SKIP_EMAIL = {
    "noreply", "no-reply", "example", "test", "wordpress",
    "sentry", "privacy", "abuse", "postmaster",
}
WA_HREF_RE = re.compile(r'wa\.me/(\+?[\d]+)|whatsapp\.com/send\?phone=([\d+]+)', re.I)
WA_TEXT_RE = re.compile(r'whatsapp[\s:]*[\+]?[\d\s\-()]{9,18}', re.I)


def _get_html(url, timeout=10):
    try:
        r = requests.get(url, headers=BROWSER_HEADERS, timeout=timeout, allow_redirects=True)
        if r.status_code == 200:
            return r.text
    except Exception:
        pass
    return ""


def scrape_contacts(url):
    """Return (email, whatsapp) by scraping the business website."""
    if not url or any(d in url.lower() for d in SOCIAL_DOMAINS):
        return "", ""

    email    = ""
    whatsapp = ""
    pages    = [url] + [urljoin(url, s) for s in ("/contact", "/contact-us", "/about")]

    for page_url in pages:
        if email and whatsapp:
            break
        html = _get_html(page_url)
        if not html:
            continue
        soup = BeautifulSoup(html, "lxml")

        # Email — prefer mailto links
        if not email:
            for tag in soup.find_all("a", href=re.compile(r'^mailto:', re.I)):
                m = re.search(r'mailto:([^\?&\s]+)', tag["href"])
                if m:
                    c = m.group(1).strip().lower()
                    if not any(s in c for s in SKIP_EMAIL):
                        email = c
                        break
        if not email:
            for c in EMAIL_RE.findall(html):
                if not any(s in c.lower() for s in SKIP_EMAIL):
                    email = c.lower()
                    break

        # WhatsApp — prefer wa.me links
        if not whatsapp:
            for tag in soup.find_all("a", href=True):
                m = WA_HREF_RE.search(tag["href"])
                if m:
                    whatsapp = (m.group(1) or m.group(2)).strip()
                    break
        if not whatsapp:
            m = WA_TEXT_RE.search(html)
            if m:
                digits = re.findall(r'[\d+]{10,13}', m.group())
                if digits:
                    whatsapp = digits[0]

    return email, whatsapp

# =========================================================================
# Excel output
# =========================================================================

COLS = [
    ("Business Name",        30),
    ("Phone Number",         18),
    ("Email Address",        32),
    ("WhatsApp Number",      18),
    ("Website Status",       16),
    ("Website / Social URL", 42),
    ("Address",              45),
    ("Rating",               10),
    ("Reviews",              10),
    ("Notes",                36),
    ("Distance (miles)",     16),
]

FILL_HEADER   = PatternFill("solid", fgColor="1A2035")
FILL_NONE     = PatternFill("solid", fgColor="FFD6D6")   # red   – no website
FILL_SOCIAL   = PatternFill("solid", fgColor="D6EAFF")   # blue  – social only
FILL_OUTDATED = PatternFill("solid", fgColor="FFF2CC")   # amber – outdated
FILL_ERROR    = PatternFill("solid", fgColor="E8E8E8")   # grey  – error
FILL_ALT      = PatternFill("solid", fgColor="F7F7F7")

STATUS_ORDER = {"none": 0, "social_only": 1, "outdated": 2, "error": 3, "active": 4}

THIN = Border(
    **{s: Side(style="thin", color="CCCCCC")
       for s in ("left", "right", "top", "bottom")}
)


def _row_fill(status, idx):
    return {
        "none":        FILL_NONE,
        "social_only": FILL_SOCIAL,
        "outdated":    FILL_OUTDATED,
        "error":       FILL_ERROR,
    }.get(status, FILL_ALT if idx % 2 == 0 else PatternFill(fill_type=None))


def save_spreadsheet(leads):
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"wigan_hair_leads_{ts}.xlsx"
    wb       = openpyxl.Workbook()

    # ── Leads sheet ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Leads"

    for col, (hdr, width) in enumerate(COLS, 1):
        c = ws.cell(1, col, hdr)
        c.fill      = FILL_HEADER
        c.font      = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = THIN
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 32

    sorted_leads = sorted(
        leads,
        key=lambda b: (STATUS_ORDER.get(b["status"], 9), b["distance_mi"]),
    )

    for ri, biz in enumerate(sorted_leads, 2):
        fill   = _row_fill(biz["status"], ri)
        values = [
            biz["name"],
            biz["phone"],
            biz["email"],
            biz["whatsapp"],
            biz["status"].upper().replace("_", " "),
            biz["website"],
            biz["address"],
            biz["rating"],
            biz["reviews"],
            biz["notes"],
            biz["distance_mi"],
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(ri, col, val)
            c.fill      = fill
            c.alignment = Alignment(vertical="center")
            c.border    = THIN
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Summary sheet ─────────────────────────────────────────────────────
    ws2  = wb.create_sheet("Summary")
    cnts = {s: sum(1 for b in leads if b["status"] == s)
            for s in ("none", "social_only", "outdated", "error")}
    rows = [
        ("Search area",          f"{RADIUS_MI}-mile radius of Wigan, UK"),
        ("Date generated",       datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Data sources",         "OpenStreetMap (Overpass API) + Yell.com"),
        ("",                     ""),
        ("Total leads",          len(leads)),
        ("No website",           cnts["none"]),
        ("Social media only",    cnts["social_only"]),
        ("Outdated website",     cnts["outdated"]),
        ("Site unreachable",     cnts["error"]),
        ("",                     ""),
        ("Have email",           sum(1 for b in leads if b["email"])),
        ("Have WhatsApp",        sum(1 for b in leads if b["whatsapp"])),
        ("Have phone",           sum(1 for b in leads if b["phone"])),
        ("",                     ""),
        ("Colour key",           ""),
        ("RED  = No website",    "Best leads – no online presence at all"),
        ("BLUE = Social only",   "Only a Facebook / Instagram page"),
        ("AMBER = Outdated",     "Site exists but old or broken"),
        ("GREY = Unreachable",   "Site down or returning an error"),
    ]
    for r, (label, val) in enumerate(rows, 1):
        ws2.cell(r, 1, label).font = Font(bold=True, size=10)
        ws2.cell(r, 2, val)
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 42

    wb.save(filename)
    return filename, sorted_leads

# =========================================================================
# Main
# =========================================================================

def main():
    print()
    print("╔══════════════════════════════════════════════════════════════╗")
    print("║   Barber & Hairdresser Lead Finder – 50 miles of Wigan      ║")
    print("║   Data: OpenStreetMap + Yell.com  |  No API key needed      ║")
    print("╚══════════════════════════════════════════════════════════════╝")
    print()

    # ── 1. Fetch from both sources ────────────────────────────────────────
    print("[1/3] Fetching businesses …\n")
    osm_biz  = fetch_from_overpass()
    print()
    yell_biz = fetch_from_yell()

    raw = deduplicate(osm_biz, yell_biz)
    print(f"\n  {len(raw)} unique businesses "
          f"({len(osm_biz)} from OSM + {len(yell_biz)} from Yell, after dedup).")

    if not raw:
        print("\n  No results found. Check your internet connection and try again.")
        input("  Press Enter to exit …")
        sys.exit(1)

    # ── 2. Check websites & scrape contacts ───────────────────────────────
    print(f"\n[2/3] Checking websites and scraping contact info …\n")
    leads = []

    for idx, biz in enumerate(raw, 1):
        name    = biz["name"]
        website = biz["website"]
        print(f"  [{idx}/{len(raw)}] {name}")

        status, notes = check_website(website)
        print(f"         → {status}  {notes}")

        if status == "active":
            continue   # skip businesses with a current, working website

        email, whatsapp = "", ""
        if website and status != "none":
            email, whatsapp = scrape_contacts(website)
            if email:    print(f"         email:    {email}")
            if whatsapp: print(f"         whatsapp: {whatsapp}")

        leads.append({
            **biz,
            "status":   status,
            "notes":    notes,
            "email":    email,
            "whatsapp": whatsapp,
        })

    # ── 3. Save spreadsheet ───────────────────────────────────────────────
    print(f"\n[3/3] Saving {len(leads)} leads to spreadsheet …")
    filename, sorted_leads = save_spreadsheet(leads)

    cnts = {s: sum(1 for b in sorted_leads if b["status"] == s)
            for s in ("none", "social_only", "outdated", "error")}

    print()
    print("╔══════════════════════════════════════════════════════════════╗")
    print(f"║  Done!  File: {filename:<47}║")
    print("╠══════════════════════════════════════════════════════════════╣")
    print(f"║  Total leads        : {len(sorted_leads):<38}║")
    print(f"║  No website (red)   : {cnts['none']:<38}║")
    print(f"║  Social only (blue) : {cnts['social_only']:<38}║")
    print(f"║  Outdated (amber)   : {cnts['outdated']:<38}║")
    print(f"║  With email found   : {sum(1 for b in sorted_leads if b['email']):<38}║")
    print(f"║  With WhatsApp      : {sum(1 for b in sorted_leads if b['whatsapp']):<38}║")
    print("╚══════════════════════════════════════════════════════════════╝")
    print()
    input("  Press Enter to exit …")


if __name__ == "__main__":
    main()
